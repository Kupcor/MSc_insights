import hyper_parameters as hp
import torch.optim as optim
import torch

import numpy as np

"""
Function: Select Optimizer
    Param:
        model - model to train
        opt_arg - selected optimizer, e.g.: "Adam"
        lr - learning rate
    Function initialize selected optimizer and returns its instance
"""
def select_optimizer(model, opt_arg=hp.optimizer_arg, lr=hp.lr):
    if opt_arg == "Adam":
        optimizer = optim.Adam(model.parameters(), lr)
    elif opt_arg == "Adagrad":
        optimizer = optim.Adagrad(model.parameters(), lr)
    elif opt_arg == "RMSprop":
        optimizer = optim.RMSprop(model.parameters(), lr)
    elif opt_arg == "AdamW":
        optimizer = optim.AdamW(model.parameters(), lr)
    # Not aplicable for now
    #elif opt_arg == "LBFGS":
    #    optimizer = optim.LBFGS(model.parameters(), lr, max_iter = 50)
    elif opt_arg == "SGD":
        optimizer = optim.SGD(model.parameters(), lr)
    elif opt_arg == "Nadam":
        optimizer = optim.NAdam(model.parameters(), lr)
    else:
        optimizer = optim.Adam(model.parameters(), lr)
    return optimizer

"""
Function: Save Model
    Param:
        model - model to save
        hidden_layers_neurons - string of numbers of neurons in hidden layers, e.g.: 10-10-10 - 3 hidden layers, 10 neurons in each
        learning_rate - model learning rate
        numb_epochs - number of lerning epochs
        optimizer - optimizer that was used during training
        test_loss - test loss during model validation - MSE - average squared difference between the predicted values and the actual values
        r2 - R-Squared - coefficient of determination - regression score function
    Function saves trained model
"""
def save_model(model, hidden_layers_neurons, learning_rate, num_epochs, optimizer, test_loss, r2, date=hp.today):
    neurons_str = '-'.join(map(str, hidden_layers_neurons))
    file_name = "model_{}_hidden_layers_{}_{}_{}_{}_{}_testLoss__{}_R2_{}".format(
    len(hidden_layers_neurons),
    neurons_str,
    learning_rate,
    num_epochs,
    optimizer.__class__.__name__,
    date,
    test_loss,
    r2
    )
    save_path = "trained_models/" + file_name + '.pth'
    torch.save(model.state_dict(), save_path)

"""
Function: Save Results To Excel
    Param:
        predictions - model outputs
        file_name - xlsx file to load
        sheet_name - new sheet name
        combination - currenty combination (predictions on a given chemical composition)
        time - time series array
        ground_truth - true values of preditions
    Function overwrites provided excel file and adds model results (results and diagrams) to it
    Need to be used with combination of create_file_with_unique_sets in bulk_predictions function in Ann_wraopper.py
"""
def save_results_to_excel(predictions, file_name, sheet_name, combination, time, ground_truth, r2, prediciton_loss, mae, huber_loss_value):
    from openpyxl import load_workbook
    import openpyxl
    import matplotlib.pyplot as plt

    workbook = load_workbook(file_name)

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    sheet.cell(row=1, column=12, value="Mass change - predictions")

    for i, prediction in enumerate(predictions, start=2):
        sheet.cell(row=i, column=12, value=prediction)

    fig, ax = plt.subplots()
    ax.scatter(time, predictions, label='Predictions')
    ax.scatter(time, ground_truth, label='Ground Truth')

    ax.set_xlabel('Time')
    ax.set_ylabel('Mass Change')
    ax.set_title(f"{combination} | {r2} | MSE Loss {prediciton_loss:.4f} | MAE {mae:.4f}\nLoss: {huber_loss_value}")
    ax.legend()
    ax.grid()

    img_path = f'image/{sheet_name}.png' 
    plt.savefig(img_path)

    img = openpyxl.drawing.image.Image(img_path)
    sheet.add_image(img, 'M10')

    fig2, ax_2 = plt.subplots()
    ax_2.scatter(ground_truth, predictions, label='Predictions')

    z = np.polyfit(ground_truth, predictions, 1)
    p = np.poly1d(z)
    plt.plot(ground_truth, p(ground_truth), 'r--', label='Trend Line')

    ax_2.set_xlabel('Mass change - ground truth')
    ax_2.set_ylabel('Mass Change - predictions')
    ax_2.set_title(f"{combination} | {r2} | MSE Loss {prediciton_loss:.4f} | MAE {mae:.4f}")
    ax_2.legend()
    ax_2.grid()
    img_path_2 = f'image/{sheet_name}_2.png' 
    plt.savefig(img_path_2)
    img_2 = openpyxl.drawing.image.Image(img_path_2)
    sheet.add_image(img_2, 'W10')

    workbook.save(file_name)

"""
Function save results to excel new data
Param:
        predictions - model outputs
        file_name - xlsx file to load
        sheet_name - new sheet name
        combination - currenty combination (predictions on a given chemical composition)
        time - time series array
    Function overwrites provided excel file and adds model results (results and diagrams) to it
    Need to be used with combination of create_file_with_unique_sets in bulk_predictions_on_new_data function in Ann_wraopper.py
"""
def save_results_to_excel_new_data(predictions, file_name, sheet_name, combination, time):
    from openpyxl import load_workbook
    import openpyxl
    import matplotlib.pyplot as plt

    workbook = load_workbook(file_name)

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    sheet.cell(row=1, column=12, value="Mass change - predictions")

    for i, prediction in enumerate(predictions, start=2):
        sheet.cell(row=i, column=12, value=prediction)

    fig, ax = plt.subplots()
    ax.scatter(time, predictions, label='Predictions', s=10)

    ax.set_xlabel('Time')
    ax.set_ylabel('Mass Change')
    ax.set_title(combination)
    ax.legend()
    ax.grid()

    img_path = f'image/{sheet_name}.png' 
    plt.savefig(img_path)

    img = openpyxl.drawing.image.Image(img_path)
    sheet.add_image(img, 'M10') 
    workbook.save(file_name)

"""
Create file with unique sets
Read basic data file and categorize data to unique sets | training and tests data
"""
def create_file_with_unique_sets(file_name='bulk_results'):
    import pandas as pd
    import matplotlib.pyplot as plt

    df = pd.read_excel('data/data_cleared_with_old_data.xlsx')

    selected_row = df[['Temperature [C]', 'Mo [at%]', 'Nb [at%]', 'Ta [at%]', 'Ti [at%]', 'Cr [at%]', 'Al [at%]', 'W [at%]', 'Zr [at%]']]

    unique = selected_row.drop_duplicates()
    with pd.ExcelWriter(f'{file_name}', engine='xlsxwriter') as writer:
        iterator = 1
        for i, row in unique.iterrows():
            sheet_name = f"combination_{iterator}"

            combination = ', '.join(row.iloc[:-1].astype(str))

            subset = df[(df['Temperature [C]'] == row['Temperature [C]']) & 
                        (df['Mo [at%]'] == row['Mo [at%]']) & 
                        (df['Nb [at%]'] == row['Nb [at%]']) & 
                        (df['Ta [at%]'] == row['Ta [at%]']) & 
                        (df['Ti [at%]'] == row['Ti [at%]']) & 
                        (df['Cr [at%]'] == row['Cr [at%]']) & 
                        (df['Al [at%]'] == row['Al [at%]']) & 
                        (df['W [at%]'] == row['W [at%]']) & 
                        (df['Zr [at%]'] == row['Zr [at%]'])]
            
            #subset = subset.drop(columns=['Parametr X'])

            subset['Mass Change [mg.cm2]'] = pd.to_numeric(subset['Mass Change [mg.cm2]'], errors='coerce')

            subset.dropna(subset=['Mass Change [mg.cm2]'], inplace=True)
            subset.dropna(how='all', inplace=True)

            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            iterator += 1

            worksheet = writer.sheets[sheet_name]
            chart = writer.book.add_chart({'type': 'scatter'})
            chart.add_series({
                'name': "Ground truth",
                'categories': f"='{sheet_name}'!$J$2:$J${len(subset) + 1}",
                'values': f"='{sheet_name}'!$K$2:$K${len(subset) + 1}",
                'marker': {'type': 'circle', 'size': 2},
            })
            chart.set_title({'name': combination})
            chart.set_x_axis({'name': 'Time'})
            chart.set_y_axis({'name': 'Mass Change'})
            worksheet.insert_chart('N2', chart)

"""
Create sets of new data to predict
Create a xlsx with same combinations of temp and composition and linear time
"""
def create_sets_of_new_data_to_predict(file_name):
    import pandas as pd
    from openpyxl import Workbook

    data = pd.read_excel("data/temp.xlsx")

    wb = Workbook()

    for index, row in data.iterrows():
        ws = wb.create_sheet(title=f"Arkusz_{index}")

        headers = list(data.columns)
        ws.append(headers[:9] + ["Time [h]"] + headers[9:])

        time = [round(x * 0.1, 1) for x in range(0, 241)]

        for t in time:
            row_with_time = list(row[:9]) + [t] + list(row[9:])
            ws.append(row_with_time)

    default_sheet = wb.active
    wb.remove(default_sheet)

    wb.save(f"{file_name}")


"""
Function select file
    Functions is used to select previously saved model and return its path and file name in order to use that parameters in other functions
"""
def select_file():
    import tkinter as tk
    from tkinter import filedialog
    
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(initialdir=r"C:\Users\Piotr Kupczyk\Mój folder\Studia\Informatyczna techniczna\Praca magisterska\Model\trained_models")
    if file_path:
        file_name = file_path.split("/")[-1]
    else:
        return None
    
    print(f"{file_path}")
    return file_path, file_name
